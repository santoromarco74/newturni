#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
GESTIONE TURNI NEGOZIO
Programma per la pianificazione dei turni settimanali/mensili del personale
Scritto in Python con interfaccia a menu interattivo
"""

from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
import json
import calendar
from typing import List, Dict, Set, Tuple
import random


class Addetto:
    """Classe che rappresenta un addetto/dipendente"""

    def __init__(self, nome: str, ore_contratto: int, ore_max_settimanale: int, straordinario: bool):
        """
        Inizializza un addetto

        Args:
            nome: Nome dell'addetto
            ore_contratto: Ore settimanali previste dal contratto (MINIMO)
            ore_max_settimanale: Ore massime settimanali che può lavorare
            straordinario: Se può fare straordinario (True/False)
        """
        self.nome = nome
        self.ore_contratto = ore_contratto  # MINIMO settimanale
        self.ore_max_settimanale = ore_max_settimanale  # MASSIMO settimanale
        self.straordinario = straordinario
        # Traccia ore realizzate per settimana: {numero_settimana: ore}
        self.ore_per_settimana = {}
        self.giorni_riposo = set()  # giorni della settimana (0=lunedì, 6=domenica)
        self.ferie_permessi = []  # lista di date
        self.turni_assegnati = {}  # {data: turno}

    def aggiungi_giorno_riposo(self, giorno: int):
        """Aggiunge un giorno di riposo settimanale (0=lunedì, 6=domenica)"""
        self.giorni_riposo.add(giorno)

    def rimuovi_giorno_riposo(self, giorno: int):
        """Rimuove un giorno di riposo settimanale"""
        if giorno in self.giorni_riposo:
            self.giorni_riposo.remove(giorno)

    def aggiungi_ferie(self, data: datetime):
        """Aggiunge una data di ferie/permesso"""
        if data not in self.ferie_permessi:
            self.ferie_permessi.append(data)

    def rimuovi_ferie(self, data: datetime):
        """Rimuove una data di ferie/permesso"""
        if data in self.ferie_permessi:
            self.ferie_permessi.remove(data)

    def puo_lavorare(self, data: datetime) -> bool:
        """Verifica se l'addetto può lavorare in una data specifica"""
        # Controlla se è un giorno di riposo settimanale
        if data.weekday() in self.giorni_riposo:
            return False

        # Controlla se è in ferie/permesso
        if data in self.ferie_permessi:
            return False

        return True

    def get_ore_settimana(self, numero_settimana: int) -> float:
        """Restituisce le ore realizzate in una settimana specifica"""
        return self.ore_per_settimana.get(numero_settimana, 0)

    def add_ore_settimana(self, numero_settimana: int, ore: float):
        """Aggiunge ore realizzate in una settimana specifica"""
        if numero_settimana not in self.ore_per_settimana:
            self.ore_per_settimana[numero_settimana] = 0
        self.ore_per_settimana[numero_settimana] += ore

    def puo_aggiungere_ore_settimana(self, numero_settimana: int, ore: float) -> bool:
        """Verifica se può aggiungere altre ore in una settimana senza superare il massimo"""
        ore_attuali = self.get_ore_settimana(numero_settimana)
        return (ore_attuali + ore) <= self.ore_max_settimanale

    def __repr__(self):
        return f"Addetto({self.nome}, {self.ore_contratto}h min, max {self.ore_max_settimanale}h/settimana)"


class Turno:
    """Classe che rappresenta un turno"""

    def __init__(self, nome: str, ora_inizio: str, ora_fine: str):
        """
        Inizializza un turno

        Args:
            nome: Nome del turno (es. 'Mattina', 'Pomeriggio')
            ora_inizio: Ora di inizio (formato HH:MM)
            ora_fine: Ora di fine (formato HH:MM)
        """
        self.nome = nome
        self.ora_inizio = ora_inizio
        self.ora_fine = ora_fine
        self.ore = self._calcola_ore()

    def _calcola_ore(self) -> int:
        """Calcola le ore del turno"""
        # Converte le ore in formato 24h per il calcolo
        h_inizio, m_inizio = map(int, self.ora_inizio.split(':'))
        h_fine, m_fine = map(int, self.ora_fine.split(':'))

        ore_inizio = h_inizio + m_inizio / 60
        ore_fine = h_fine + m_fine / 60

        if ore_fine < ore_inizio:
            ore_fine += 24  # Se il turno va oltre la mezzanotte

        return ore_fine - ore_inizio

    def __repr__(self):
        return f"Turno({self.nome}, {self.ora_inizio}-{self.ora_fine}, {self.ore}h)"


class TurnoManager:
    """Classe principale per la gestione della pianificazione dei turni"""

    # Giorni festivi (non lavorativi)
    GIORNI_FESTIVI = [
        (1, 1),      # 1 gennaio
        (4, 20),     # 20 aprile
        (5, 1),      # 1 maggio
        (12, 25),    # 25 dicembre
        (12, 26),    # 26 dicembre
    ]

    def __init__(self):
        self.addetti: List[Addetto] = []
        self.turni: List[Turno] = []
        self.mese = datetime.now().month
        self.anno = datetime.now().year
        self.pianificazione = {}  # {data: {addetto: turno}}

    def is_festivo(self, data: datetime) -> bool:
        """Verifica se una data è festiva"""
        return (data.month, data.day) in self.GIORNI_FESTIVI

    def is_domenica(self, data: datetime) -> bool:
        """Verifica se una data è domenica"""
        return data.weekday() == 6

    def get_giorni_mese(self) -> List[datetime]:
        """Restituisce tutti i giorni lavorativi del mese"""
        _, num_giorni = calendar.monthrange(self.anno, self.mese)
        giorni = []

        for giorno in range(1, num_giorni + 1):
            data = datetime(self.anno, self.mese, giorno)
            # Esclude giorni festivi
            if not self.is_festivo(data):
                giorni.append(data)

        return giorni

    def get_settimane_mese(self) -> Dict[int, List[datetime]]:
        """Restituisce le settimane del mese con i loro giorni"""
        # Una settimana inizia lunedì (0) e finisce domenica (6)
        giorni = self.get_giorni_mese()
        settimane = {}
        settimana_attuale = None

        for data in giorni:
            # Numero della settimana ISO (1-53)
            num_settimana = data.isocalendar()[1]

            if num_settimana not in settimane:
                settimane[num_settimana] = []

            settimane[num_settimana].append(data)

        return settimane

    def get_numero_settimana(self, data: datetime) -> int:
        """Restituisce il numero della settimana ISO per una data"""
        return data.isocalendar()[1]

    def aggiungi_addetto(self, addetto: Addetto):
        """Aggiunge un addetto alla lista"""
        if addetto not in self.addetti:
            self.addetti.append(addetto)

    def rimuovi_addetto(self, nome: str):
        """Rimuove un addetto dalla lista"""
        self.addetti = [a for a in self.addetti if a.nome != nome]

    def aggiungi_turno(self, turno: Turno):
        """Aggiunge un tipo di turno disponibile"""
        if turno not in self.turni:
            self.turni.append(turno)

    def rimuovi_turno(self, nome: str):
        """Rimuove un tipo di turno"""
        self.turni = [t for t in self.turni if t.nome != nome]

    def pianifica_turni(self):
        """
        Algoritmo principale di pianificazione dei turni.
        Assegna i turni cercando di:
        - Alternare i turni il più possibile
        - Rispettare le ore settimanali minime e massime
        """
        giorni = self.get_giorni_mese()

        if not self.addetti or not self.turni:
            print("Errore: Aggiungere almeno un addetto e un turno")
            return False

        # Reset dei turni assegnati
        for addetto in self.addetti:
            addetto.turni_assegnati = {}
            addetto.ore_per_settimana = {}

        self.pianificazione = {giorno: {} for giorno in giorni}

        # Algoritmo di assegnazione per ogni giorno
        for data in giorni:
            num_settimana = self.get_numero_settimana(data)

            # Filtra addetti disponibili
            disponibili = [a for a in self.addetti if a.puo_lavorare(data)]

            if not disponibili:
                continue

            # Ordina per criterio: meno ore nella settimana, ultimo turno diverso
            disponibili.sort(
                key=lambda a: (
                    a.get_ore_settimana(num_settimana),
                    self._get_priorita_turno(a, data)
                )
            )

            # Assegna turni a rotazione, rispettando il massimo settimanale
            turni_da_assegnare = self._seleziona_turni(disponibili)

            for addetto, turno in zip(disponibili, turni_da_assegnare):
                # Controlla se può aggiungere questo turno senza superare il massimo settimanale
                if addetto.puo_aggiungere_ore_settimana(num_settimana, turno.ore):
                    self.pianificazione[data][addetto.nome] = turno
                    addetto.turni_assegnati[data] = turno
                    addetto.add_ore_settimana(num_settimana, turno.ore)

        return True

    def _get_priorita_turno(self, addetto: Addetto, data: datetime) -> int:
        """
        Calcola la priorità del turno per evitare turni uguali consecutivi.
        Restituisce 0 se il turno è diverso da quello della giorno precedente.
        """
        data_precedente = data - timedelta(days=1)

        if data_precedente not in addetto.turni_assegnati:
            return 0

        ultimo_turno = addetto.turni_assegnati[data_precedente]

        # Se ha turni assegnati e l'ultimo è diverso da quelli disponibili, ritorna 0
        for turno in self.turni:
            if turno != ultimo_turno:
                return 0

        return 1

    def _seleziona_turni(self, addetti: List[Addetto]) -> List[Turno]:
        """
        Seleziona i turni da assegnare agli addetti disponibili.
        Tenta di diversificare evitando ripetizioni.
        """
        turni_disponibili = self.turni.copy()
        turni_selezionati = []

        for addetto in addetti:
            if turni_disponibili:
                # Preferisce turni diversi da quelli recenti dell'addetto
                turno = self._scegli_turno_migliore(addetto, turni_disponibili)
                turni_selezionati.append(turno)
                turni_disponibili.remove(turno)
            else:
                # Se finiscono i turni, ricomincia da capo
                turni_disponibili = self.turni.copy()
                if turni_disponibili:
                    turno = turni_disponibili.pop(0)
                    turni_selezionati.append(turno)

        return turni_selezionati

    def _scegli_turno_migliore(self, addetto: Addetto, turni: List[Turno]) -> Turno:
        """
        Sceglie il turno migliore per un addetto tra quelli disponibili.
        Preferisce turni diversi dall'ultimo assegnato.
        """
        if not turni:
            return self.turni[0]

        # Se non ha ultimi turni, sceglie il primo disponibile
        if not addetto.turni_assegnati:
            return turni[0]

        # Prende l'ultimo turno assegnato
        ultimi_turni = list(addetto.turni_assegnati.values())
        if ultimi_turni:
            ultimo_turno = ultimi_turni[-1]

            # Preferisce un turno diverso dall'ultimo
            for turno in turni:
                if turno != ultimo_turno:
                    return turno

        return turni[0]

    def genera_statistiche(self) -> Dict:
        """Genera statistiche sulla pianificazione"""
        stats = {
            'ore_totali_per_addetto': {},
            'ore_per_settimana': {},  # {nome_addetto: {num_settimana: ore}}
            'giorni_lavorati_per_addetto': {},
            'domeniche_lavorate': 0,
            'dettaglio_domeniche': {}
        }

        for addetto in self.addetti:
            # Calcola ore totali sommando le ore per settimana
            ore_totali = sum(addetto.ore_per_settimana.values())
            stats['ore_totali_per_addetto'][addetto.nome] = ore_totali
            stats['ore_per_settimana'][addetto.nome] = addetto.ore_per_settimana.copy()
            stats['giorni_lavorati_per_addetto'][addetto.nome] = len(addetto.turni_assegnati)

        # Conta domeniche lavorate
        for data, assegnazioni in self.pianificazione.items():
            if assegnazioni:  # Se ci sono assegnazioni quel giorno
                if self.is_domenica(data):
                    stats['domeniche_lavorate'] += len(assegnazioni)
                    for nome_addetto in assegnazioni.keys():
                        if nome_addetto not in stats['dettaglio_domeniche']:
                            stats['dettaglio_domeniche'][nome_addetto] = 0
                        stats['dettaglio_domeniche'][nome_addetto] += 1

        return stats

    def esporta_excel(self, nome_file: str = None) -> str:
        """
        Esporta la pianificazione in file Excel

        Args:
            nome_file: Nome del file di output (default: turni_MMMM_AAAA.xlsx)

        Returns:
            Percorso del file creato
        """
        if nome_file is None:
            nome_file = f"turni_{self.mese:02d}_{self.anno}.xlsx"

        # Crea workbook
        wb = Workbook()

        # --- FOGLIO 1: Pianificazione ---
        ws_pianificazione = wb.active
        ws_pianificazione.title = "Pianificazione"

        # Intestazione
        giorni = self.get_giorni_mese()
        ws_pianificazione['A1'] = f"PIANIFICAZIONE TURNI - {self._nome_mese()} {self.anno}"
        ws_pianificazione['A1'].font = Font(bold=True, size=14)
        ws_pianificazione.merge_cells('A1:D1')

        # Intestazioni colonne
        ws_pianificazione['A3'] = "Data"
        ws_pianificazione['B3'] = "Giorno"
        ws_pianificazione['C3'] = "Addetto"
        ws_pianificazione['D3'] = "Turno (Orario)"

        for col in ['A', 'B', 'C', 'D']:
            ws_pianificazione[f'{col}3'].font = Font(bold=True, color="FFFFFF")
            ws_pianificazione[f'{col}3'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

        # Dati - una riga per ogni assegnazione (addetto + turno)
        row = 4
        for data in giorni:
            assegnazioni = self.pianificazione.get(data, {})

            # Colora festivi e domeniche
            if self.is_festivo(data):
                colore = "FFCCCC"  # Rosso chiaro
            elif self.is_domenica(data):
                colore = "FFFFCC"  # Giallo chiaro
            else:
                colore = "FFFFFF"  # Bianco

            if assegnazioni:
                # Una riga per ogni addetto assegnato a quella data
                for nome_addetto, turno in assegnazioni.items():
                    ws_pianificazione[f'A{row}'] = data.strftime("%d/%m/%Y")
                    ws_pianificazione[f'B{row}'] = self._nome_giorno_italiano(data.weekday())
                    ws_pianificazione[f'C{row}'] = nome_addetto
                    ws_pianificazione[f'D{row}'] = f"{turno.nome} ({turno.ora_inizio}-{turno.ora_fine})"

                    # Applica colori
                    for col in ['A', 'B', 'C', 'D']:
                        ws_pianificazione[f'{col}{row}'].fill = PatternFill(start_color=colore, end_color=colore, fill_type="solid")

                    row += 1
            else:
                # Nessun addetto assegnato quel giorno
                ws_pianificazione[f'A{row}'] = data.strftime("%d/%m/%Y")
                ws_pianificazione[f'B{row}'] = self._nome_giorno_italiano(data.weekday())
                ws_pianificazione[f'C{row}'] = "-"
                ws_pianificazione[f'D{row}'] = "Nessun turno"

                # Applica colori
                for col in ['A', 'B', 'C', 'D']:
                    ws_pianificazione[f'{col}{row}'].fill = PatternFill(start_color=colore, end_color=colore, fill_type="solid")

                row += 1

        # Autofit colonne
        ws_pianificazione.column_dimensions['A'].width = 12
        ws_pianificazione.column_dimensions['B'].width = 15
        ws_pianificazione.column_dimensions['C'].width = 18
        ws_pianificazione.column_dimensions['D'].width = 25

        # --- FOGLIO 2: Statistiche ---
        ws_stats = wb.create_sheet("Statistiche")

        stats = self.genera_statistiche()

        ws_stats['A1'] = "STATISTICHE PIANIFICAZIONE"
        ws_stats['A1'].font = Font(bold=True, size=14)

        # Ore totali per addetto
        ws_stats['A3'] = "ORE TOTALI PER ADDETTO"
        ws_stats['A3'].font = Font(bold=True)
        ws_stats['A4'] = "Addetto"
        ws_stats['B4'] = "Ore"

        row = 5
        for nome, ore in stats['ore_totali_per_addetto'].items():
            ws_stats[f'A{row}'] = nome
            ws_stats[f'B{row}'] = ore
            row += 1

        # Giorni lavorati
        ws_stats['D3'] = "GIORNI LAVORATI"
        ws_stats['D3'].font = Font(bold=True)
        ws_stats['D4'] = "Addetto"
        ws_stats['E4'] = "Giorni"

        row = 5
        for nome, giorni_lav in stats['giorni_lavorati_per_addetto'].items():
            ws_stats[f'D{row}'] = nome
            ws_stats[f'E{row}'] = giorni_lav
            row += 1

        # Domeniche lavorate
        ws_stats['A13'] = "DOMENICHE LAVORATE"
        ws_stats['A13'].font = Font(bold=True)
        ws_stats['A14'] = "Addetto"
        ws_stats['B14'] = "Giorni"

        row = 15
        for nome, giorni in stats['dettaglio_domeniche'].items():
            ws_stats[f'A{row}'] = nome
            ws_stats[f'B{row}'] = giorni
            row += 1

        ws_stats.column_dimensions['A'].width = 20
        ws_stats.column_dimensions['B'].width = 15
        ws_stats.column_dimensions['D'].width = 20
        ws_stats.column_dimensions['E'].width = 15

        # --- FOGLIO 3: Dettagli Addetti ---
        ws_addetti = wb.create_sheet("Dettagli Addetti")

        ws_addetti['A1'] = "DETTAGLI ADDETTI E VINCOLI"
        ws_addetti['A1'].font = Font(bold=True, size=14)

        ws_addetti['A3'] = "Nome"
        ws_addetti['B3'] = "Ore Contratto (min)"
        ws_addetti['C3'] = "Ore Max (sett)"
        ws_addetti['D3'] = "Straordinario"
        ws_addetti['E3'] = "Giorni Riposo"
        ws_addetti['F3'] = "Giorni Ferie"

        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            ws_addetti[f'{col}3'].font = Font(bold=True, color="FFFFFF")
            ws_addetti[f'{col}3'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

        row = 4
        for addetto in self.addetti:
            ws_addetti[f'A{row}'] = addetto.nome
            ws_addetti[f'B{row}'] = addetto.ore_contratto
            ws_addetti[f'C{row}'] = addetto.ore_max_settimanale
            ws_addetti[f'D{row}'] = "Sì" if addetto.straordinario else "No"

            giorni_riposo = [self._nome_giorno_italiano(g) for g in sorted(addetto.giorni_riposo)]
            ws_addetti[f'E{row}'] = ", ".join(giorni_riposo) if giorni_riposo else "-"

            ferie = [f.strftime("%d/%m") for f in sorted(addetto.ferie_permessi)]
            ws_addetti[f'F{row}'] = ", ".join(ferie) if ferie else "-"

            row += 1

        ws_addetti.column_dimensions['A'].width = 20
        ws_addetti.column_dimensions['B'].width = 15
        ws_addetti.column_dimensions['C'].width = 12
        ws_addetti.column_dimensions['D'].width = 15
        ws_addetti.column_dimensions['E'].width = 20
        ws_addetti.column_dimensions['F'].width = 20

        # Salva workbook
        wb.save(nome_file)

        return os.path.abspath(nome_file)

    def _nome_mese(self) -> str:
        """Restituisce il nome del mese in italiano"""
        nomi = [
            "Gennaio", "Febbraio", "Marzo", "Aprile", "Maggio", "Giugno",
            "Luglio", "Agosto", "Settembre", "Ottobre", "Novembre", "Dicembre"
        ]
        return nomi[self.mese - 1]

    def _nome_giorno_italiano(self, weekday: int) -> str:
        """Converte il numero del giorno della settimana in nome italiano"""
        nomi = ["Lunedì", "Martedì", "Mercoledì", "Giovedì", "Venerdì", "Sabato", "Domenica"]
        return nomi[weekday]


class MenuInterattivo:
    """Classe per gestire il menu interattivo"""

    def __init__(self):
        self.manager = TurnoManager()
        self.running = True

    def mostra_menu_principale(self):
        """Mostra il menu principale"""
        print("\n" + "="*60)
        print("   GESTIONE TURNI NEGOZIO".center(60))
        print("="*60)
        print("\n1. Gestione Addetti")
        print("2. Gestione Turni")
        print("3. Pianificazione Turni")
        print("4. Visualizza Pianificazione")
        print("5. Statistiche")
        print("6. Esporta in Excel")
        print("7. Impostazioni Mese/Anno")
        print("0. Esci")
        print("\n" + "-"*60)

    def menu_principale(self):
        """Gestisce il menu principale"""
        while self.running:
            self.mostra_menu_principale()
            scelta = input("Seleziona un'opzione: ").strip()

            if scelta == '1':
                self.menu_addetti()
            elif scelta == '2':
                self.menu_turni()
            elif scelta == '3':
                self.pianifica_turni()
            elif scelta == '4':
                self.visualizza_pianificazione()
            elif scelta == '5':
                self.mostra_statistiche()
            elif scelta == '6':
                self.esporta_excel()
            elif scelta == '0':
                print("\nGrazie per aver utilizzato il programma!")
                self.running = False
            else:
                print("Opzione non valida, riprova.")

    def menu_addetti(self):
        """Menu per gestione addetti"""
        while True:
            print("\n" + "="*60)
            print("   GESTIONE ADDETTI".center(60))
            print("="*60)
            print("\n1. Aggiungi Addetto")
            print("2. Visualizza Addetti")
            print("3. Modifica Addetto")
            print("4. Rimuovi Addetto")
            print("5. Torna al Menu Principale")
            print("\n" + "-"*60)

            scelta = input("Seleziona un'opzione: ").strip()

            if scelta == '1':
                self.aggiungi_addetto()
            elif scelta == '2':
                self.visualizza_addetti()
            elif scelta == '3':
                self.modifica_addetto()
            elif scelta == '4':
                self.rimuovi_addetto()
            elif scelta == '5':
                break
            else:
                print("Opzione non valida.")

    def aggiungi_addetto(self):
        """Aggiunge un nuovo addetto"""
        print("\n--- Aggiungi Nuovo Addetto ---")

        nome = input("Nome addetto: ").strip()

        if any(a.nome.lower() == nome.lower() for a in self.manager.addetti):
            print("Addetto già esistente!")
            return

        try:
            ore_contratto = int(input("Ore settimanali di contratto (MINIMO): "))
            ore_max_settimanale = int(input("Ore massime settimanali (MASSIMO): "))

            straordinario_input = input("Può fare straordinario? (s/n): ").strip().lower()
            straordinario = straordinario_input == 's'

            addetto = Addetto(nome, ore_contratto, ore_max_settimanale, straordinario)

            # Menu giorni di riposo
            print("\n--- Seleziona Giorni di Riposo Settimanale ---")
            giorni = ["Lunedì", "Martedì", "Mercoledì", "Giovedì", "Venerdì", "Sabato", "Domenica"]

            for i, giorno in enumerate(giorni):
                scelta = input(f"Riposo il {giorno}? (s/n): ").strip().lower()
                if scelta == 's':
                    addetto.aggiungi_giorno_riposo(i)

            self.manager.aggiungi_addetto(addetto)
            print(f"\n✓ Addetto '{nome}' aggiunto con successo!")

            # Opzione per aggiungere ferie
            self.aggiungi_ferie_addetto(addetto)

        except ValueError:
            print("Errore: Inserisci valori numerici corretti.")

    def aggiungi_ferie_addetto(self, addetto: Addetto):
        """Aggiunge ferie/permessi a un addetto"""
        print(f"\n--- Ferie/Permessi per {addetto.nome} ---")

        while True:
            data_input = input("Inserisci data di ferie (gg/mm) o premi Invio per terminare: ").strip()

            if not data_input:
                break

            try:
                giorno, mese = map(int, data_input.split('/'))
                data = datetime(self.manager.anno, mese, giorno)
                addetto.aggiungi_ferie(data)
                print(f"✓ Ferie aggiunta: {data.strftime('%d/%m/%Y')}")
            except (ValueError, IndexError):
                print("Formato non valido. Usa gg/mm")
            except Exception as e:
                print(f"Errore: {e}")

    def visualizza_addetti(self):
        """Visualizza l'elenco degli addetti"""
        print("\n--- Elenco Addetti ---")

        if not self.manager.addetti:
            print("Nessun addetto inserito.")
            return

        for i, addetto in enumerate(self.manager.addetti, 1):
            giorni_riposo = [["Lun", "Mar", "Mer", "Gio", "Ven", "Sab", "Dom"][g]
                             for g in sorted(addetto.giorni_riposo)]

            print(f"\n{i}. {addetto.nome}")
            print(f"   Ore Contratto (min): {addetto.ore_contratto}h/settimana")
            print(f"   Ore Max (sett): {addetto.ore_max_settimanale}h/settimana")
            print(f"   Straordinario: {'Sì' if addetto.straordinario else 'No'}")
            print(f"   Giorni Riposo: {', '.join(giorni_riposo) if giorni_riposo else 'Nessuno'}")
            print(f"   Ferie: {len(addetto.ferie_permessi)} giorni")

    def modifica_addetto(self):
        """Modifica i dati di un addetto"""
        self.visualizza_addetti()

        nome = input("\nNome dell'addetto da modificare: ").strip()

        addetto = next((a for a in self.manager.addetti if a.nome.lower() == nome.lower()), None)

        if not addetto:
            print("Addetto non trovato.")
            return

        print(f"\n--- Modifica {addetto.nome} ---")
        print("1. Aggiungi Giorno di Riposo")
        print("2. Rimuovi Giorno di Riposo")
        print("3. Aggiungi Ferie")
        print("4. Rimuovi Ferie")
        print("5. Modifica Ore Max")
        print("6. Torna Indietro")

        scelta = input("Seleziona: ").strip()

        if scelta == '1':
            giorni = ["Lunedì", "Martedì", "Mercoledì", "Giovedì", "Venerdì", "Sabato", "Domenica"]
            for i, giorno in enumerate(giorni):
                print(f"{i}: {giorno}")
            try:
                giorno_idx = int(input("Seleziona giorno (numero): "))
                addetto.aggiungi_giorno_riposo(giorno_idx)
                print("✓ Giorno di riposo aggiunto")
            except (ValueError, IndexError):
                print("Numero non valido")

        elif scelta == '2':
            giorni = ["Lunedì", "Martedì", "Mercoledì", "Giovedì", "Venerdì", "Sabato", "Domenica"]
            giorni_riposo = [giorni[g] for g in sorted(addetto.giorni_riposo)]
            print(f"Giorni di riposo attuali: {', '.join(giorni_riposo)}")

            try:
                giorno_idx = int(input("Numero del giorno da rimuovere: "))
                addetto.rimuovi_giorno_riposo(giorno_idx)
                print("✓ Giorno di riposo rimosso")
            except ValueError:
                print("Numero non valido")

        elif scelta == '3':
            self.aggiungi_ferie_addetto(addetto)

        elif scelta == '4':
            if not addetto.ferie_permessi:
                print("Nessuna feria registrata.")
                return

            ferie_str = [f.strftime("%d/%m/%Y") for f in addetto.ferie_permessi]
            for i, feria in enumerate(ferie_str, 1):
                print(f"{i}. {feria}")

            try:
                scelta_feria = int(input("Seleziona numero feria da rimuovere: ")) - 1
                if 0 <= scelta_feria < len(addetto.ferie_permessi):
                    feria = addetto.ferie_permessi[scelta_feria]
                    addetto.rimuovi_ferie(feria)
                    print("✓ Feria rimossa")
            except ValueError:
                print("Numero non valido")

        elif scelta == '5':
            try:
                nuove_ore = int(input("Nuove ore massime settimanali: "))
                addetto.ore_max_settimanale = nuove_ore
                print("✓ Ore max settimanali modificate")
            except ValueError:
                print("Numero non valido")

    def rimuovi_addetto(self):
        """Rimuove un addetto"""
        self.visualizza_addetti()

        nome = input("\nNome dell'addetto da rimuovere: ").strip()

        if not any(a.nome.lower() == nome.lower() for a in self.manager.addetti):
            print("Addetto non trovato.")
            return

        conferma = input(f"Confermi rimozione di '{nome}'? (s/n): ").strip().lower()

        if conferma == 's':
            self.manager.rimuovi_addetto(nome)
            print(f"✓ Addetto '{nome}' rimosso.")

    def menu_turni(self):
        """Menu per gestione turni"""
        while True:
            print("\n" + "="*60)
            print("   GESTIONE TURNI".center(60))
            print("="*60)
            print("\n1. Aggiungi Turno")
            print("2. Visualizza Turni")
            print("3. Rimuovi Turno")
            print("4. Torna al Menu Principale")
            print("\n" + "-"*60)

            scelta = input("Seleziona un'opzione: ").strip()

            if scelta == '1':
                self.aggiungi_turno()
            elif scelta == '2':
                self.visualizza_turni()
            elif scelta == '3':
                self.rimuovi_turno()
            elif scelta == '4':
                break
            else:
                print("Opzione non valida.")

    def aggiungi_turno(self):
        """Aggiunge un nuovo turno"""
        print("\n--- Aggiungi Nuovo Turno ---")

        nome = input("Nome turno (es. Mattina, Pomeriggio): ").strip()

        if any(t.nome.lower() == nome.lower() for t in self.manager.turni):
            print("Turno già esistente!")
            return

        ora_inizio = input("Ora inizio (HH:MM): ").strip()
        ora_fine = input("Ora fine (HH:MM): ").strip()

        try:
            # Valida formato HH:MM
            datetime.strptime(ora_inizio, "%H:%M")
            datetime.strptime(ora_fine, "%H:%M")

            turno = Turno(nome, ora_inizio, ora_fine)
            self.manager.aggiungi_turno(turno)
            print(f"\n✓ Turno '{nome}' aggiunto ({turno.ore}h)")
        except ValueError:
            print("Formato orario non valido. Usa HH:MM")

    def visualizza_turni(self):
        """Visualizza i turni disponibili"""
        print("\n--- Turni Disponibili ---")

        if not self.manager.turni:
            print("Nessun turno inserito.")
            return

        for i, turno in enumerate(self.manager.turni, 1):
            print(f"{i}. {turno.nome}: {turno.ora_inizio}-{turno.ora_fine} ({turno.ore}h)")

    def rimuovi_turno(self):
        """Rimuove un turno"""
        self.visualizza_turni()

        nome = input("\nNome turno da rimuovere: ").strip()

        if not any(t.nome.lower() == nome.lower() for t in self.manager.turni):
            print("Turno non trovato.")
            return

        conferma = input(f"Confermi rimozione di '{nome}'? (s/n): ").strip().lower()

        if conferma == 's':
            self.manager.rimuovi_turno(nome)
            print(f"✓ Turno '{nome}' rimosso.")

    def pianifica_turni(self):
        """Avvia l'algoritmo di pianificazione"""
        print("\n" + "="*60)
        print("   PIANIFICAZIONE TURNI".center(60))
        print("="*60)

        if not self.manager.addetti:
            print("Errore: Aggiungere almeno un addetto.")
            return

        if not self.manager.turni:
            print("Errore: Aggiungere almeno un turno.")
            return

        print(f"\nPianificazione per: {self.manager._nome_mese()} {self.manager.anno}")
        print(f"Addetti: {len(self.manager.addetti)}")
        print(f"Turni disponibili: {len(self.manager.turni)}")

        conferma = input("\nAvviare la pianificazione? (s/n): ").strip().lower()

        if conferma == 's':
            if self.manager.pianifica_turni():
                print("\n✓ Pianificazione completata con successo!")
            else:
                print("\n✗ Errore durante la pianificazione.")

    def visualizza_pianificazione(self):
        """Visualizza la pianificazione corrente"""
        if not self.manager.pianificazione:
            print("\nNessuna pianificazione disponibile. Esegui prima la pianificazione.")
            return

        print("\n" + "="*60)
        print(f"   PIANIFICAZIONE {self.manager._nome_mese()} {self.manager.anno}".center(60))
        print("="*60 + "\n")

        giorni = self.manager.get_giorni_mese()

        # Stampa in formato tabella semplificato
        for data in giorni:
            assegnazioni = self.manager.pianificazione.get(data, {})
            giorno_settimana = self.manager._nome_giorno_italiano(data.weekday())
            data_str = data.strftime("%d/%m")

            tipo = ""
            if self.manager.is_festivo(data):
                tipo = "[FESTIVO]"
            elif self.manager.is_domenica(data):
                tipo = "[DOMENICA]"

            if assegnazioni:
                addetti_str = ", ".join([f"{nome} ({turno.nome})" for nome, turno in assegnazioni.items()])
                print(f"{data_str} {giorno_settimana:12} {tipo:12} → {addetti_str}")
            else:
                print(f"{data_str} {giorno_settimana:12} {tipo:12} → Nessun turno")

    def mostra_statistiche(self):
        """Visualizza le statistiche della pianificazione"""
        if not self.manager.pianificazione:
            print("\nNessuna pianificazione disponibile.")
            return

        stats = self.manager.genera_statistiche()

        print("\n" + "="*60)
        print("   STATISTICHE PIANIFICAZIONE".center(60))
        print("="*60)

        print("\n--- ORE TOTALI PER ADDETTO (mese) ---")
        for nome, ore in sorted(stats['ore_totali_per_addetto'].items()):
            print(f"{nome:20} {ore:5.0f}h totali nel mese")

        print("\n--- ORE PER SETTIMANA ---")
        for nome, ore_settimane in sorted(stats['ore_per_settimana'].items()):
            addetto = next(a for a in self.manager.addetti if a.nome == nome)
            if ore_settimane:
                dettagli = ", ".join([f"Sett {s}: {o:.0f}h" for s, o in sorted(ore_settimane.items())])
                print(f"{nome:20} {dettagli}")
                media = sum(ore_settimane.values()) / len(ore_settimane)
                print(f"{'':20} Media settimanale: {media:.1f}h (contratto: {addetto.ore_contratto}h min, max {addetto.ore_max_settimanale}h)")

        print("\n--- GIORNI LAVORATI PER ADDETTO ---")
        for nome, giorni in sorted(stats['giorni_lavorati_per_addetto'].items()):
            print(f"{nome:20} {giorni:3} giorni")

        if stats['dettaglio_domeniche']:
            print("\n--- DOMENICHE LAVORATE ---")
            for nome, giorni in sorted(stats['dettaglio_domeniche'].items()):
                print(f"{nome:20} {giorni:3} domeniche")

        print("\n--- RIEPILOGO ---")
        print(f"Mese: {self.manager._nome_mese()} {self.manager.anno}")
        print(f"Giorni lavorativi: {len(self.manager.get_giorni_mese())}")
        print(f"Addetti totali: {len(self.manager.addetti)}")

    def esporta_excel(self):
        """Esporta la pianificazione in Excel"""
        if not self.manager.pianificazione:
            print("\nNessuna pianificazione disponibile.")
            return

        print("\n" + "="*60)
        print("   ESPORTAZIONE EXCEL".center(60))
        print("="*60)

        nome_file = input("\nNome file (senza estensione, default: turni_MMMM_AAAA): ").strip()

        if nome_file:
            nome_file = f"{nome_file}.xlsx"

        try:
            percorso = self.manager.esporta_excel(nome_file)
            print(f"\n✓ File esportato con successo!")
            print(f"Percorso: {percorso}")
        except Exception as e:
            print(f"\n✗ Errore durante l'esportazione: {e}")

    def impostazioni_mese(self):
        """Modifica il mese e l'anno di pianificazione"""
        print("\n--- Impostazioni Mese/Anno ---")

        try:
            mese = int(input(f"Mese (1-12, default {self.manager.mese}): ") or self.manager.mese)
            anno = int(input(f"Anno (default {self.manager.anno}): ") or self.manager.anno)

            if 1 <= mese <= 12 and anno >= 2000:
                self.manager.mese = mese
                self.manager.anno = anno
                print(f"✓ Data impostata: {self.manager._nome_mese()} {anno}")
            else:
                print("Valori non validi.")
        except ValueError:
            print("Inserisci numeri validi.")


# === PUNTO DI ENTRATA DEL PROGRAMMA ===
if __name__ == "__main__":
    try:
        menu = MenuInterattivo()
        menu.menu_principale()
    except KeyboardInterrupt:
        print("\n\nProgramma interrotto dall'utente.")
    except Exception as e:
        print(f"\nErrore non gestito: {e}")
        import traceback
        traceback.print_exc()
